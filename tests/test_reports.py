"""Admin reports: status buckets, missing items, filters, exports (Phase 8)."""

import pytest
from django.urls import reverse

from academics.models import Course, Term
from folders.models import FolderStatus, ItemStatus
from reports import services


@pytest.fixture
def terms(db):
    return {
        "fall": Term.objects.create(season=Term.Season.FALL, year=2026, is_current=True),
        "spring": Term.objects.create(season=Term.Season.SPRING, year=2026),
    }


def _course(term, instructor, code, program="BSCS", section="A"):
    return Course.objects.create(
        code=code, title=code, section=section, program=program,
        study_semester=1, instructor=instructor, term=term,
    )


def _set_status(course, status):
    course.folder.status = status
    course.folder.save(update_fields=["status"])


def _make_all_required_available(folder):
    folder.items.filter(is_required=True).update(status=ItemStatus.AVAILABLE)


# --- Task 1: report content ------------------------------------------------

@pytest.mark.django_db
def test_status_buckets(terms, faculty_user):
    c1 = _course(terms["fall"], faculty_user, "CS101")
    c2 = _course(terms["fall"], faculty_user, "CS102")
    c3 = _course(terms["fall"], faculty_user, "CS103")
    _set_status(c1, FolderStatus.CERTIFIED)
    _set_status(c2, FolderStatus.MID_SUBMITTED)
    _set_status(c3, FolderStatus.DRAFT)

    rows = {r["course"].code: r for r in services.get_report()}
    assert rows["CS101"]["report_status"] == services.CERTIFIED
    assert rows["CS102"]["report_status"] == services.IN_REVIEW
    assert rows["CS103"]["report_status"] == services.PENDING


@pytest.mark.django_db
def test_missing_required_items_listed(terms, faculty_user):
    course = _course(terms["fall"], faculty_user, "CS101")
    folder = course.folder
    _make_all_required_available(folder)
    # Knock one required item back to pending.
    item = folder.items.filter(is_required=True).first()
    item.status = ItemStatus.PENDING
    item.save(update_fields=["status"])

    row = services.get_report()[0]
    assert row["missing_count"] == 1
    assert item.title in row["missing"]


@pytest.mark.django_db
def test_na_items_not_counted_as_missing(terms, faculty_user):
    course = _course(terms["fall"], faculty_user, "CS101")
    folder = course.folder
    _make_all_required_available(folder)
    item = folder.items.filter(is_required=True).first()
    item.status = ItemStatus.NOT_APPLICABLE
    item.save(update_fields=["status"])

    row = services.get_report()[0]
    assert row["missing_count"] == 0


@pytest.mark.django_db
def test_summary_counts(terms, faculty_user):
    _set_status(_course(terms["fall"], faculty_user, "CS101"), FolderStatus.CERTIFIED)
    _set_status(_course(terms["fall"], faculty_user, "CS102"), FolderStatus.CERTIFIED)
    _set_status(_course(terms["fall"], faculty_user, "CS103"), FolderStatus.FINAL_SUBMITTED)
    summary = services.summarise(services.get_report())
    assert summary["counts"]["certified"] == 2
    assert summary["counts"]["in_review"] == 1
    assert summary["total"] == 3


def test_report_is_admin_only(faculty_client):
    assert faculty_client.get(reverse("report")).status_code == 403


@pytest.mark.django_db
def test_report_page_renders(admin_client, terms, faculty_user):
    _set_status(_course(terms["fall"], faculty_user, "CS101"), FolderStatus.CERTIFIED)
    resp = admin_client.get(reverse("report"))
    assert resp.status_code == 200
    assert b"CS101" in resp.content
    assert b"Certified" in resp.content


# --- Task 2: filters -------------------------------------------------------

@pytest.mark.django_db
def test_filter_by_term(admin_client, terms, faculty_user):
    _course(terms["fall"], faculty_user, "CS101")
    _course(terms["spring"], faculty_user, "CS102")
    resp = admin_client.get(reverse("report"), {"term": terms["fall"].pk})
    assert b"CS101" in resp.content
    assert b"CS102" not in resp.content


@pytest.mark.django_db
def test_filter_by_status(admin_client, terms, faculty_user):
    _set_status(_course(terms["fall"], faculty_user, "CS101"), FolderStatus.CERTIFIED)
    _set_status(_course(terms["fall"], faculty_user, "CS102"), FolderStatus.DRAFT)
    resp = admin_client.get(reverse("report"), {"status": "certified"})
    assert b"CS101" in resp.content
    assert b"CS102" not in resp.content


@pytest.mark.django_db
def test_filter_by_program_and_faculty(admin_client, terms, faculty_user):
    from django.contrib.auth import get_user_model
    other = get_user_model().objects.create_user(email="t2@uiit.edu.pk", name="Teacher Two")
    _course(terms["fall"], faculty_user, "CS101", program="BSCS")
    _course(terms["fall"], other, "SE201", program="BSSE")

    resp = admin_client.get(reverse("report"), {"program": "BSSE"})
    assert b"SE201" in resp.content and b"CS101" not in resp.content

    resp = admin_client.get(reverse("report"), {"faculty": faculty_user.pk})
    assert b"CS101" in resp.content and b"SE201" not in resp.content


# --- Task 3: exports -------------------------------------------------------

@pytest.mark.django_db
def test_pdf_export(admin_client, terms, faculty_user):
    _course(terms["fall"], faculty_user, "CS101")
    resp = admin_client.get(reverse("report_export_pdf"))
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"
    assert "attachment" in resp["Content-Disposition"]


@pytest.mark.django_db
def test_xlsx_export(admin_client, terms, faculty_user):
    from io import BytesIO
    from openpyxl import load_workbook

    _course(terms["fall"], faculty_user, "CS101")
    resp = admin_client.get(reverse("report_export_xlsx"))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]

    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Course"
    assert ws.cell(row=2, column=1).value == "CS101"


@pytest.mark.django_db
def test_export_respects_filters(admin_client, terms, faculty_user):
    from io import BytesIO
    from openpyxl import load_workbook

    _set_status(_course(terms["fall"], faculty_user, "CS101"), FolderStatus.CERTIFIED)
    _set_status(_course(terms["fall"], faculty_user, "CS102"), FolderStatus.DRAFT)
    resp = admin_client.get(reverse("report_export_xlsx"), {"status": "certified"})
    wb = load_workbook(BytesIO(resp.content))
    codes = [c[0].value for c in wb.active.iter_rows(min_row=2)]
    assert "CS101" in codes and "CS102" not in codes


def test_exports_are_admin_only(faculty_client):
    assert faculty_client.get(reverse("report_export_pdf")).status_code == 403
    assert faculty_client.get(reverse("report_export_xlsx")).status_code == 403
